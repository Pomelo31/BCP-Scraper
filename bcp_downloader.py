#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Descargador específico para archivos Excel del Banco Central del Paraguay
Versión mejorada con búsqueda más precisa y manejo de errores robusto
"""

import requests
from bs4 import BeautifulSoup
import os
import time
import logging
import re
from urllib.parse import urljoin, urlparse
from datetime import datetime
from openpyxl import load_workbook
import json
import csv
import unicodedata

# Importar requests-html si está disponible
try:
    from requests_html import HTMLSession
    REQUESTS_HTML_AVAILABLE = True
except ImportError:
    REQUESTS_HTML_AVAILABLE = False
    print("WARNING: requests-html no esta disponible. Instalala con: pip install requests-html")

# Intentar importar cloudscraper si esta disponible
try:
    import cloudscraper
    CLOUDSCRAPER_AVAILABLE = True
except ImportError:
    CLOUDSCRAPER_AVAILABLE = False
    print("WARNING: cloudscraper not available. Install with: pip install cloudscraper")

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bcp_downloader.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class BCPDownloader:
    """Descargador especializado para archivos del Banco Central del Paraguay"""
    

    def __init__(self):
        self.base_url = "https://www.bcp.gov.py"
        self.target_url = "https://www.bcp.gov.py/web/institucional/boletines-formato-macros"

        self.default_headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'es-ES,es;q=0.9,en-US;q=0.8,en;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br, zstd',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Cache-Control': 'max-age=0',
            'DNT': '1'
        }

        self.download_headers = {
            'User-Agent': self.default_headers['User-Agent'],
            'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel,application/octet-stream,*/*',
            'Accept-Language': self.default_headers['Accept-Language'],
            'Accept-Encoding': self.default_headers['Accept-Encoding'],
            'Connection': 'keep-alive',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'Referer': self.target_url,
            'Cache-Control': 'no-cache',
            'Pragma': 'no-cache'
        }

        self.session = None
        self.session_type = None
        self.max_page_retries = 3
        self.max_download_retries = 2
        self._init_session()

        # URLs de ejemplo como fallback (Julio 2025)
        self.fallback_urls = {
            'tabla_bancos': 'https://www.bcp.gov.py/documents/20117/0/1.1+Tablas+Bolet%C3%ADn+Bancos+Jul25+%282%29+1.xlsx/c44c7087-46a1-4569-14be-32b782902f79?t=1756399940286',
            'tabla_financieras': 'https://www.bcp.gov.py/documents/20117/0/2.1+Tablas+Bolet%C3%ADn+Financieras+Jul25+%281%29+1.xlsx/64887684-ae2a-9000-e3ea-dd936598c6ef?t=1756399892013'
        }

        # Archivos objetivo a buscar
        self.target_files = {
            'tabla_bancos': ['tabla de bancos', 'bancos', 'sistema bancario'],
            'tabla_financieras': ['tabla de financieras', 'financieras', 'entidades financieras']
        }

    def _init_session(self, prefer_cloudscraper=True):
        if prefer_cloudscraper and CLOUDSCRAPER_AVAILABLE:
            try:
                self.session = cloudscraper.create_scraper(
                    browser={'browser': 'chrome', 'platform': 'windows', 'mobile': False}
                )
                self.session_type = 'cloudscraper'
                logger.info("Sesion HTTP inicializada con cloudscraper")
            except Exception as exc:
                logger.warning(f"No se pudo crear sesion con cloudscraper: {exc}")
                self.session = requests.Session()
                self.session_type = 'requests'
        else:
            if prefer_cloudscraper and not CLOUDSCRAPER_AVAILABLE:
                logger.info("cloudscraper no esta disponible; se usara requests.Session")
            self.session = requests.Session()
            self.session_type = 'requests'

        self._apply_default_headers()

    def _apply_default_headers(self):
        self.session.headers.clear()
        self.session.headers.update(self.default_headers)

    def _reset_session(self, reason, prefer_cloudscraper=True):
        logger.info(f"Reiniciando la sesion HTTP ({reason})")
        try:
            self.session.close()
        except Exception:
            pass
        self._init_session(prefer_cloudscraper=prefer_cloudscraper)
    

    def _get_download_basename(self, category):
        mapping = {
            'tabla_bancos': 'tabla_de_bancos',
            'tabla_financieras': 'tabla_de_financieras'
        }
        return mapping.get(category, category)

    def _get_category_suffix(self, category):
        suffix_map = {
            'tabla_bancos': 'BI',
            'tabla_financieras': 'FI'
        }
        return suffix_map.get(category, 'XX')

    def _normalize_text(self, value):
        if not value:
            return ''
        normalized = unicodedata.normalize('NFKD', value)
        ascii_text = normalized.encode('ascii', 'ignore').decode()
        ascii_text = re.sub(r'[^a-z0-9]+', ' ', ascii_text.lower())
        return ascii_text.strip()

    def extract_sheets_to_csv(self, filepath, category):
        """Extrae hojas especificas del Excel y las guarda como CSV"""
        sheet_targets = [
            {'label': 'EEFF', 'prefix': 'ER', 'keywords': ['eeff']},
            {'label': 'TC', 'prefix': 'TC', 'keywords': ['tc']},
            {'label': 'Credito Sector', 'prefix': 'CS', 'keywords': ['cred', 'sector']}
        ]

        suffix = self._get_category_suffix(category)
        output_dir = os.path.dirname(filepath)

        try:
            workbook = load_workbook(filepath, read_only=True, data_only=True)
        except Exception as exc:
            logger.error(f"Error al abrir {filepath} para extraer CSV: {exc}")
            return

        normalized_titles = []
        for worksheet in workbook.worksheets:
            normalized_titles.append((worksheet, self._normalize_text(worksheet.title)))

        try:
            for target in sheet_targets:
                target_sheet = None
                for worksheet, normalized in normalized_titles:
                    if all(keyword in normalized for keyword in target['keywords']):
                        target_sheet = worksheet
                        break

                if not target_sheet:
                    logger.warning(f"No se encontro una hoja relacionada con '{target['label']}' en {os.path.basename(filepath)}")
                    continue

                output_filename = f"{target['prefix']}{suffix}.csv"
                output_path = os.path.join(output_dir, output_filename)

                try:
                    with open(output_path, 'w', newline='', encoding='utf-8') as csv_file:
                        writer = csv.writer(csv_file)
                        for row in target_sheet.iter_rows(values_only=True):
                            writer.writerow(['' if value is None else value for value in row])
                    logger.info(f"CSV generado: {output_path}")
                except Exception as exc:
                    logger.error(f"Error al escribir CSV '{output_filename}': {exc}")
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def establish_session(self, retries=3):
        """Establece una sesion valida visitando primero la pagina principal"""
        main_page_url = self.base_url

        for attempt in range(retries):
            try:
                logger.info(f"Estableciendo sesion con el BCP (intento {attempt + 1}/{retries})")
                response = self.session.get(main_page_url, timeout=30)
            except requests.RequestException as exc:
                logger.warning(f"No se pudo establecer la sesion (intento {attempt + 1}): {exc}")
                if attempt < retries - 1:
                    if CLOUDSCRAPER_AVAILABLE:
                        self._reset_session("error al contactar la pagina principal", prefer_cloudscraper=True)
                    time.sleep(1 + attempt)
                    continue
                return False

            if response.status_code == 200:
                logger.info("Sesion establecida correctamente")
                return True

            if response.status_code == 403:
                logger.warning("Respuesta 403 al establecer la sesion")
                if attempt < retries - 1 and CLOUDSCRAPER_AVAILABLE:
                    if self.session_type != 'cloudscraper':
                        logger.info("Activando cloudscraper para resolver Cloudflare")
                    else:
                        logger.info("Renovando la sesion de cloudscraper")
                    self._reset_session("respuesta 403 en la pagina principal", prefer_cloudscraper=True)
                    time.sleep(1 + attempt)
                    continue
                return False

            logger.warning(f"Respuesta inesperada de la pagina principal: {response.status_code}")
            if attempt < retries - 1:
                time.sleep(1 + attempt)
                continue
            return False

        return False
    

    def get_page_content_with_js(self, url):
        """Obtiene el contenido HTML usando requests-html para manejar JavaScript"""
        if not REQUESTS_HTML_AVAILABLE:
            logger.warning("requests-html no esta disponible")
            return None

        session = HTMLSession()
        try:
            logger.info("Usando requests-html para manejar JavaScript...")
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
                'Accept-Language': 'es-ES,es;q=0.9,en-US;q=0.8,en;q=0.7',
                'Accept-Encoding': 'gzip, deflate, br, zstd',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            })

            response = session.get(url, timeout=30)

            if response.html:
                response.html.render(timeout=20)

            if session.cookies:
                self.session.cookies.update(session.cookies)

            class MockResponse:
                def __init__(self, html_response):
                    self.status_code = html_response.status_code
                    self.text = html_response.html.html
                    self.headers = html_response.headers
                    self.url = html_response.url

            mock_response = MockResponse(response)
            logger.info(f"Pagina cargada con requests-html ({len(mock_response.text):,} caracteres)")
            return mock_response

        except Exception as exc:
            logger.error(f"Error con requests-html: {exc}")
            return None
        finally:
            session.close()
    

    def get_page_content(self, url, max_retries=None):
        """Obtiene el contenido HTML con reintentos"""
        if max_retries is None:
            max_retries = self.max_page_retries

        for attempt in range(max_retries):
            try:
                logger.info(f"Accediendo a: {url} (intento {attempt + 1}/{max_retries})")
                response = self.session.get(url, timeout=30)
            except requests.RequestException as exc:
                logger.warning(f"Error al solicitar la pagina (intento {attempt + 1}): {exc}")
                if attempt < max_retries - 1:
                    if CLOUDSCRAPER_AVAILABLE:
                        self._reset_session("error de red al obtener la pagina", prefer_cloudscraper=True)
                        self.establish_session()
                    time.sleep(2 ** attempt)
                    continue
                logger.error(f"No se pudo obtener la pagina despues de {max_retries} intentos")
                return None

            if response.status_code == 403:
                logger.warning(f"Proteccion de Cloudflare (403) al acceder a {url}")
                if attempt < max_retries - 1 and CLOUDSCRAPER_AVAILABLE:
                    self._reset_session("403 al obtener la pagina objetivo", prefer_cloudscraper=True)
                    self.establish_session()
                    time.sleep(2 ** attempt)
                    continue
                logger.error("La proteccion de Cloudflare bloqueo la solicitud")
                return None

            try:
                response.raise_for_status()
            except requests.HTTPError as exc:
                logger.warning(f"Respuesta invalida ({exc})")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
                    continue
                return None

            content_type = response.headers.get('content-type', '').lower()
            if 'text/html' not in content_type:
                logger.warning(f"Tipo de contenido inesperado: {content_type}")

            return response

        return None
    
    def find_excel_links(self, soup):
        """Busca enlaces a archivos Excel basandose en la estructura especifica del BCP"""
        excel_links = []
        seen_urls = set()

        def add_link(data):
            url = data.get('url')
            if not url or url in seen_urls:
                return
            seen_urls.add(url)
            excel_links.append(data)

        logger.info("Buscando secciones especificas usando la estructura de la lista de documentos")
        section_items = soup.select("div.list_item.section-item")
        for item in section_items:
            title_el = item.select_one(".item_title")
            link_el = item.select_one("div.item_links a[href]")
            if not link_el:
                continue

            href = link_el.get('href', '')
            if '.xls' not in href.lower():
                continue

            title_text = title_el.get_text(strip=True) if title_el else ''
            link_text = title_text or link_el.get_text(strip=True) or 'descargar'
            file_type = self._determine_file_type_from_link(link_text, href)

            add_link({
                'url': urljoin(self.base_url, href),
                'text': link_text,
                'type': file_type,
                'method': 'structured_section'
            })

        logger.info("Buscando secciones especificas: 'Tabla de Bancos' y 'Tabla de Financieras'")
        possible_headers = soup.find_all(
            ['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'div', 'span'],
            string=re.compile(r'tabla de (bancos|financieras)', re.I)
        )

        for header in possible_headers:
            header_text = header.get_text().strip().lower()
            logger.info(f"Encontrado header: '{header_text}'")

            file_type = None
            if 'tabla de bancos' in header_text:
                file_type = 'tabla_bancos'
            elif 'tabla de financieras' in header_text:
                file_type = 'tabla_financieras'

            if file_type:
                logger.info(f"Procesando seccion: {file_type}")
                download_link = self._find_download_button_near_header(header, file_type)
                if download_link:
                    add_link(download_link)

        if len(excel_links) < 2:
            logger.info("Busqueda alternativa: buscando todos los botones de descarga")
            all_download_buttons = soup.find_all('a', string=re.compile(r'descargar', re.I))

            for button in all_download_buttons:
                context_link = self._analyze_download_button_context(button)
                if context_link:
                    add_link(context_link)

        if len(excel_links) < 2:
            logger.info("Busqueda de respaldo: enlaces directos a Excel")
            direct_links = soup.find_all('a', href=re.compile(r'\.(xlsx|xls)', re.I))
            for link in direct_links:
                href = link.get('href', '')
                text = link.get_text().strip()
                file_type = self._determine_file_type_from_link(text, href)
                add_link({
                    'url': urljoin(self.base_url, href),
                    'text': text,
                    'type': file_type,
                    'method': 'direct_link'
                })

        logger.info(f"Total de enlaces unicos encontrados: {len(excel_links)}")
        return excel_links
    
    def _find_download_button_near_header(self, header, file_type):
        """Encuentra el botón de descarga cerca de un header específico"""
        # Buscar en el contenedor padre
        current = header.parent
        max_depth = 5  # Límite de profundidad para evitar búsquedas infinitas
        
        for depth in range(max_depth):
            if current is None:
                break
                
            # Buscar botones de descarga en este nivel
            download_buttons = current.find_all('a', string=re.compile(r'descargar', re.I))
            
            for button in download_buttons:
                href = button.get('href', '')
                if href and ('.xlsx' in href.lower() or '.xls' in href.lower()):
                    logger.info(f"Encontrado botón de descarga para {file_type}: {href}")
                    return {
                        'url': urljoin(self.base_url, href),
                        'text': f"Tabla de {file_type.replace('tabla_', '').title()}",
                        'type': file_type,
                        'method': 'section_header'
                    }
            
            current = current.parent
        
        return None
    
    def _analyze_download_button_context(self, button):
        """Analiza el contexto de un botón de descarga para determinar su tipo"""
        href = button.get('href', '')
        if not href or ('.xlsx' not in href.lower() and '.xls' not in href.lower()):
            return None
        
        # Buscar títulos o texto cercano al botón
        current = button.parent
        max_depth = 3
        
        for depth in range(max_depth):
            if current is None:
                break
            
            # Buscar texto que indique el tipo de archivo
            text_content = current.get_text().lower()
            
            if 'tabla de bancos' in text_content or 'bancos' in text_content:
                return {
                    'url': urljoin(self.base_url, href),
                    'text': "Tabla de Bancos",
                    'type': 'tabla_bancos',
                    'method': 'context_analysis'
                }
            elif 'tabla de financieras' in text_content or 'financieras' in text_content:
                return {
                    'url': urljoin(self.base_url, href),
                    'text': "Tabla de Financieras", 
                    'type': 'tabla_financieras',
                    'method': 'context_analysis'
                }
            
            current = current.parent
        
        return None
    
    def _determine_file_type_from_link(self, link_text, href):
        """Determina el tipo de archivo basado en el texto del enlace o URL"""
        text_lower = link_text.lower()
        href_lower = href.lower()
        
    def categorize_links(self, links):
        """Categoriza los enlaces segun los archivos objetivo"""
        categorized = {
            'tabla_bancos': [],
            'tabla_financieras': [],
            'otros': []
        }

        for link in links:
            explicit_type = link.get('type')
            if explicit_type in ('tabla_bancos', 'tabla_financieras'):
                categorized[explicit_type].append(link)
                continue

            text = link.get('text', '').lower()
            url = link.get('url', '').lower()

            if any(keyword in text for keyword in self.target_files['tabla_bancos']) or 'banco' in url:
                categorized['tabla_bancos'].append(link)
            elif any(keyword in text for keyword in self.target_files['tabla_financieras']) or 'financier' in url:
                categorized['tabla_financieras'].append(link)
            else:
                categorized['otros'].append(link)

        return categorized


    def download_file(self, url, filename, download_dir="descargas"):
        """Descarga un archivo con manejo robusto de errores"""
        os.makedirs(download_dir, exist_ok=True)
        logger.info(f"Descargando: {filename}")

        for attempt in range(self.max_download_retries):
            try:
                response = self.session.get(
                    url,
                    headers=self.download_headers,
                    stream=True,
                    timeout=60
                )
            except requests.RequestException as exc:
                logger.error(f"Error al descargar {url} (intento {attempt + 1}): {exc}")
                if attempt < self.max_download_retries - 1:
                    if CLOUDSCRAPER_AVAILABLE:
                        self._reset_session("error de red durante la descarga", prefer_cloudscraper=True)
                        self.establish_session()
                    time.sleep(1 + attempt)
                    continue
                return None

            if response.status_code == 403:
                logger.warning(f"Proteccion de Cloudflare (403) al descargar {url}")
                if attempt < self.max_download_retries - 1 and CLOUDSCRAPER_AVAILABLE:
                    self._reset_session("403 durante la descarga", prefer_cloudscraper=True)
                    self.establish_session()
                    time.sleep(1 + attempt)
                    continue
                return None

            try:
                response.raise_for_status()
            except requests.HTTPError as exc:
                logger.error(f"Error HTTP al descargar {url}: {exc}")
                if attempt < self.max_download_retries - 1:
                    time.sleep(1 + attempt)
                    continue
                return None

            content_type = response.headers.get('content-type', '').lower()
            if content_type:
                logger.info(f"Content-Type: {content_type}")

            content_length = response.headers.get('content-length')
            if content_length:
                try:
                    logger.info(f"Tamano del archivo: {int(content_length):,} bytes")
                except ValueError:
                    logger.info(f"Tamano del archivo: {content_length} bytes")

            parsed_url = urlparse(url)
            file_ext = os.path.splitext(parsed_url.path)[1] or '.xlsx'
            filepath = os.path.join(download_dir, f"{filename}{file_ext}")

            try:
                total_size = int(content_length) if content_length else 0
            except ValueError:
                total_size = 0
            downloaded_size = 0

            try:
                with open(filepath, 'wb') as file:
                    for chunk in response.iter_content(chunk_size=8192):
                        if not chunk:
                            continue
                        file.write(chunk)
                        downloaded_size += len(chunk)
                        if total_size > 0 and downloaded_size % (1024 * 1024) == 0:
                            progress = (downloaded_size / total_size) * 100
                            logger.info(f"Progreso: {progress:.1f}%")
            except OSError as exc:
                logger.error(f"Error al escribir el archivo {filepath}: {exc}")
                if os.path.exists(filepath):
                    os.remove(filepath)
                return None

            file_size = os.path.getsize(filepath)
            logger.info(f"Archivo guardado: {filepath} ({file_size:,} bytes)")
            return filepath

        logger.error("No se pudo descargar el archivo despues de los reintentos configurados")
        return None
    
    def save_metadata(self, downloaded_files, metadata_file="metadata.json"):
        """Guarda metadatos de los archivos descargados"""
        metadata = {
            'timestamp': datetime.now().isoformat(),
            'files': []
        }
        
        for filepath in downloaded_files:
            if os.path.exists(filepath):
                stat = os.stat(filepath)
                metadata['files'].append({
                    'filename': os.path.basename(filepath),
                    'path': filepath,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                })
        
        try:
            with open(metadata_file, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, indent=2, ensure_ascii=False)
            logger.info(f"Metadatos guardados en: {metadata_file}")
        except Exception as e:
            logger.error(f"Error al guardar metadatos: {e}")
    
    def use_fallback_urls(self):
        """Usa las URLs de ejemplo como fallback"""
        logger.info("[RETRY] Usando URLs de ejemplo como fallback...")

        downloaded_files = []

        for file_type, url in self.fallback_urls.items():
            base_name = self._get_download_basename(file_type)
            logger.info(f"[DOWNLOAD] Descargando {file_type} desde URL de ejemplo...")
            filepath = self.download_file(url, base_name)
            if filepath:
                downloaded_files.append(filepath)
                self.extract_sheets_to_csv(filepath, file_type)
            time.sleep(2)  # Pausa entre descargas

        return downloaded_files
    

    def run(self):
        """Ejecuta el proceso completo de descarga"""
        logger.info("=== Iniciando Descargador BCP ===")
        start_time = time.time()
        downloaded_files = []

        if not self.establish_session():
            logger.warning("No se pudo establecer la sesion inicial con el BCP")

        response = self.get_page_content(self.target_url)

        if not response and REQUESTS_HTML_AVAILABLE:
            logger.info("Intentando cargar la pagina con requests-html...")
            response = self.get_page_content_with_js(self.target_url)

        if not response:
            logger.warning("No se pudo acceder al contenido principal del sitio")
            logger.info("Usando URLs de ejemplo como fallback...")
            downloaded_files = self.use_fallback_urls()
        else:
            soup = BeautifulSoup(response.text, 'html.parser')
            excel_links = self.find_excel_links(soup)
            categorized = self.categorize_links(excel_links)

            target_files_found = [
                key for key in ('tabla_bancos', 'tabla_financieras') if categorized[key]
            ]

            if len(target_files_found) >= 2:
                logger.info("Archivos objetivo encontrados en la pagina")
                downloaded_files = self._download_target_files(categorized)
            else:
                logger.warning("No se encontraron suficientes archivos objetivo en la pagina")
                logger.info("Intentando con URLs de ejemplo...")
                downloaded_files = self.use_fallback_urls()

        if downloaded_files:
            self.save_metadata(downloaded_files)

        end_time = time.time()
        duration = end_time - start_time

        if downloaded_files:
            logger.info(f"Proceso completado en {duration:.2f} segundos")
            logger.info(f"Archivos descargados: {len(downloaded_files)}")
            for filepath in downloaded_files:
                logger.info(f"  - {filepath}")
            return True

        logger.error(f"No se descargo ningun archivo despues de {duration:.2f} segundos")
        return False
    
    def _download_target_files(self, categorized_links):
        """Descarga los archivos objetivo encontrados"""
        downloaded_files = []

        for category, links in categorized_links.items():
            if category == 'otros' or not links:
                continue

            link = links[0]
            base_name = self._get_download_basename(category)

            logger.info(f"[DOWNLOAD] Descargando {category}...")
            filepath = self.download_file(link['url'], base_name)
            if filepath:
                downloaded_files.append(filepath)
                self.extract_sheets_to_csv(filepath, category)

            time.sleep(2)  # Pausa entre descargas

        return downloaded_files

def main():
    """Funcion principal"""
    downloader = BCPDownloader()
    success = downloader.run()

    if success:
        print("\nDESCARGA COMPLETA: Descarga completada exitosamente!")
    else:
        print("\nERROR: La descarga fallo. Revisa los logs para mas detalles.")

if __name__ == "__main__":
    main()
