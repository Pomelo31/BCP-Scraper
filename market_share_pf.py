#!/usr/bin/env python3
"""
Pipeline para cargar los CSV generados por el downloader del BCP a DuckDB
y producir gráficos de market share y composición de cartera PF.
"""

from __future__ import annotations

import argparse
import json
import math
import textwrap
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import duckdb
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.ticker import FuncFormatter

# Rutas por defecto
DATA_DIR = Path("descargas")
OUT_DIR = Path("out")
DB_PATH = Path("data/mi_base.duckdb")
CUTOFF_DEFAULT = "2025-03-31"

# Productos objetivo dentro de las hojas de Excel
PRODUCT_LABELS = {
    "CONSUMO": "Consumo",
    "CONSUMO PERSONAS FISICAS": "Consumo",
    "VIVIENDA": "Vivienda",
}

# Clasificación desde la hoja de tarjetas
TC_CLASSIFICATION = "SALDO"

# Grupos de entidades (normalizados sin acentos)
BANK_GROUPS: Dict[str, str] = {
    "ATLAS": "ATLAS/FAMILIAR",
    "FAMILIAR": "ATLAS/FAMILIAR",
    "CONTINENTAL": "CONTINENTAL/RIO",
    "RIO": "CONTINENTAL/RIO",
    "RÍO": "CONTINENTAL/RIO",
    "R?O": "CONTINENTAL/RIO",
    "BANCOP": "BANCOP",
    "BASA": "BASA",
    "BNA": "BNA",
    "BNF": "BNF",
    "CITIBANK": "CITIBANK",
    "DO BRASIL": "DO BRASIL",
    "GNB": "GNB",
    "INTERFISA": "INTERFISA",
    "ITAÚ": "ITAU",
    "ITAU": "ITAU",
    "SOLAR": "SOLAR",
    "SUDAMERIS": "SUDAMERIS",
    "UENO": "UENO",
    "ZETA": "ZETA",
    "CEFISA": "CEFISA",
    "FIC": "FIC",
    "FINLATINA": "FINLATINA",
    "FPJ": "FPJ",
    "TU FINANCIERA": "TU FINANCIERA",
}


def ensure_out_dir() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)


def normalize_name(name: str) -> str:
    import unicodedata

    norm = (
        unicodedata.normalize("NFKD", name)
        .encode("ascii", "ignore")
        .decode("ascii")
        .upper()
        .strip()
    )
    norm = norm.replace("  ", " ")
    return norm


def map_bank(raw_name: str) -> str:
    key = normalize_name(raw_name)
    return BANK_GROUPS.get(key, key.title())


def parse_month(value) -> date:
    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)
    if isinstance(value, (int, float)):
        # Excel serial number
        base = datetime(1899, 12, 30)
        return (base + pd.to_timedelta(value, unit="D")).date().replace(day=1)
    if isinstance(value, str):
        txt = value.strip()
        if not txt:
            raise ValueError("Fecha vacía en hoja")
        txt = txt.replace("-", "/").replace(".", "/")
        for fmt in ("%Y/%m/%d", "%Y/%m", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(txt, fmt)
                return dt.date().replace(day=1)
            except ValueError:
                continue
    raise ValueError(f"No se pudo interpretar la fecha: {value!r}")


def read_credit_sector(excel_path: Path) -> pd.DataFrame:
    from openpyxl import load_workbook

    sheet_name = "5. Cred. por sector"
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    banks: List[str] = []
    records: List[dict] = []
    current_month: Optional[date] = None

    for row in ws.iter_rows(values_only=True):
        if not banks:
            if row[1] == "Fecha" and str(row[2]).startswith("Sector"):
                idx = 3
                while idx < len(row) and row[idx]:
                    banks.append(str(row[idx]))
                    idx += 1
            continue

        val_month, sector = row[1], row[2]
        if isinstance(val_month, str) and val_month.startswith("Total"):
            break
        if val_month:
            current_month = parse_month(val_month)
        if not current_month or not sector:
            continue

        sector_norm = normalize_name(str(sector))
        if sector_norm not in PRODUCT_LABELS:
            continue

        producto = PRODUCT_LABELS[sector_norm]
        for bank, value in zip(banks, row[3 : 3 + len(banks)]):
            if value is None or (isinstance(value, float) and math.isnan(value)):
                continue
            records.append(
                {
                    "fecha": current_month,
                    "banco": map_bank(bank),
                    "producto": producto,
                    "monto": float(value),
                }
            )

    wb.close()
    return pd.DataFrame(records)


def read_tarjetas(excel_path: Path) -> pd.DataFrame:
    from openpyxl import load_workbook

    sheet_name = "7. E. TC"
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    banks: List[str] = []
    records: List[dict] = []
    current_month: Optional[date] = None

    for row in ws.iter_rows(values_only=True):
        if not banks:
            if row[1] == "Fecha" and row[2]:
                idx = 3
                while idx < len(row) and row[idx]:
                    banks.append(str(row[idx]))
                    idx += 1
            continue

        val_month, clasif = row[1], row[2]
        if isinstance(val_month, str) and val_month.startswith("Total"):
            break
        if val_month:
            current_month = parse_month(val_month)
        if not current_month or not clasif:
            continue

        if normalize_name(str(clasif)) != TC_CLASSIFICATION:
            continue

        for bank, value in zip(banks, row[3 : 3 + len(banks)]):
            if value is None or (isinstance(value, float) and math.isnan(value)):
                continue
            records.append(
                {
                    "fecha": current_month,
                    "banco": map_bank(bank),
                    "producto": "TC",
                    "monto": float(value),
                }
            )

    wb.close()
    return pd.DataFrame(records)


def build_cartera_pf() -> pd.DataFrame:
    excel_bancos = DATA_DIR / "tabla_de_bancos.xlsx"
    excel_financieras = DATA_DIR / "tabla_de_financieras.xlsx"
    if not excel_bancos.exists() or not excel_financieras.exists():
        raise FileNotFoundError(
            "No se encontraron los archivos Excel en descargas/. "
            "Ejecuta bcp_downloader.py previamente."
        )

    df_parts = [
        read_credit_sector(excel_bancos),
        read_credit_sector(excel_financieras),
        read_tarjetas(excel_bancos),
        read_tarjetas(excel_financieras),
    ]

    df = pd.concat(df_parts, ignore_index=True)
    df = (
        df.groupby(["fecha", "banco", "producto"], as_index=False, sort=False)["monto"]
        .sum()
    )
    df["fecha"] = pd.to_datetime(df["fecha"])
    df["banco"] = df["banco"].str.strip()
    df["producto"] = df["producto"].str.strip()
    return df


def load_into_duckdb(df: pd.DataFrame, db_path: Path = DB_PATH) -> None:
    ensure_out_dir()
    conn = duckdb.connect(str(db_path))
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cartera_pf (
            fecha DATE,
            banco TEXT,
            producto TEXT,
            monto DOUBLE
        )
        """
    )
    conn.execute("DELETE FROM cartera_pf")
    conn.register("df_cartera", df)
    conn.execute(
        "INSERT INTO cartera_pf SELECT fecha, banco, producto, monto FROM df_cartera"
    )
    conn.unregister("df_cartera")
    conn.close()


def get_share_df(conn: duckdb.DuckDBPyConnection, cutoff: str) -> pd.DataFrame:
    sql = """
    WITH base AS (
        SELECT
            date_trunc('month', fecha) AS mes,
            banco,
            producto,
            SUM(monto) AS monto_mes
        FROM cartera_pf
        WHERE fecha BETWEEN date_trunc('month', CAST(? AS DATE)) - INTERVAL 23 MONTH
              AND date_trunc('month', CAST(? AS DATE))
        GROUP BY 1,2,3
    ),
    tot_mes AS (
        SELECT mes, SUM(monto_mes) AS total_mes
        FROM base
        GROUP BY 1
    ),
    share AS (
        SELECT
            b.mes,
            b.banco,
            b.producto,
            b.monto_mes,
            t.total_mes,
            CASE WHEN t.total_mes = 0 THEN NULL ELSE b.monto_mes / t.total_mes END AS share_mes
        FROM base b
        JOIN tot_mes t USING (mes)
    )
    SELECT * FROM share
    ORDER BY mes, banco, producto;
    """
    df = conn.execute(sql, [cutoff, cutoff]).df()
    df["mes"] = pd.to_datetime(df["mes"])
    return df


def format_pct(ax):
    ax.yaxis.set_major_formatter(FuncFormatter(lambda y, _: f"{y:.0%}"))


def plot_market_share(df_share: pd.DataFrame, cutoff: str) -> Path:
    ensure_out_dir()
    cutoff_month = pd.to_datetime(cutoff).to_period("M").to_timestamp()
    banks = sorted(df_share["banco"].unique())
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = plt.cm.get_cmap("tab20", len(banks))

    for idx, bank in enumerate(banks):
        data = df_share[df_share["banco"] == bank]
        ax.plot(
            data["mes"],
            data["share_mes"],
            label=bank,
            color=colors(idx),
            linewidth=2,
        )

    ax.axvline(cutoff_month, color="gray", linestyle="--", linewidth=1)
    ax.set_title("Market share de cartera PF por banco")
    start = df_share["mes"].min()
    ax.set_xlabel(f"Meses ({start.strftime('%b-%y')} a {cutoff_month.strftime('%b-%y')})")
    ax.set_ylabel("Participación")
    format_pct(ax)
    ax.set_ylim(0, min(1, df_share["share_mes"].max() * 1.2))
    ax.legend(bbox_to_anchor=(1.05, 1), loc="upper left")
    ax.grid(True, which="both", axis="y", linestyle=":", linewidth=0.5)
    fig.autofmt_xdate(rotation=45)
    out_path = OUT_DIR / "market_share_24m.png"
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    return out_path


def prepare_composition(df_share: pd.DataFrame, cutoff: str) -> pd.DataFrame:
    cutoff_month = pd.to_datetime(cutoff).to_period("M").to_timestamp()
    df_cutoff = df_share[df_share["mes"] == cutoff_month].copy()
    if df_cutoff.empty:
        return df_cutoff

    totals = (
        df_cutoff.groupby("banco", as_index=False)["monto_mes"]
        .sum()
        .rename(columns={"monto_mes": "total_banco"})
    )
    df_cutoff = df_cutoff.merge(totals, on="banco", how="left")
    df_cutoff["share_producto"] = np.where(
        df_cutoff["total_banco"] > 0,
        df_cutoff["monto_mes"] / df_cutoff["total_banco"],
        np.nan,
    )
    return df_cutoff


def plot_composition_cutoff(df_share: pd.DataFrame, cutoff: str) -> Path:
    ensure_out_dir()
    df_cutoff = prepare_composition(df_share, cutoff)
    if df_cutoff.empty:
        raise ValueError("No hay datos para el cutoff especificado")

    cutoff_month = pd.to_datetime(cutoff).to_period("M").to_timestamp()
    productos = sorted(df_cutoff["producto"].unique())
    bancos = sorted(df_cutoff["banco"].unique())

    fig, ax = plt.subplots(figsize=(10, 6))
    bottom = np.zeros(len(bancos))
    colors = plt.cm.get_cmap("Set2", len(productos))

    for idx, producto in enumerate(productos):
        row = (
            df_cutoff[df_cutoff["producto"] == producto]
            .set_index("banco")
            .reindex(bancos)
        )
        valores = row["share_producto"].to_numpy()
        ax.bar(
            bancos,
            valores,
            bottom=bottom,
            color=colors(idx),
            label=producto,
        )
        bottom += np.nan_to_num(valores)

    ax.set_title(
        f"Composición de cartera PF por producto ({cutoff_month.strftime('%b-%y')})"
    )
    ax.set_ylabel("Participación dentro del banco")
    ax.set_xlabel("Banco")
    ax.set_ylim(0, 1)
    format_pct(ax)
    ax.legend(title="Producto", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right")
    ax.grid(True, axis="y", linestyle=":", linewidth=0.5)
    fig.tight_layout()
    out_path = OUT_DIR / "composicion_cutoff.png"
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    return out_path


def save_csvs(df_share: pd.DataFrame, df_compo: pd.DataFrame) -> None:
    df_share.to_csv(OUT_DIR / "share_24m.csv", index=False)
    df_compo.to_csv(OUT_DIR / "compo_cutoff.csv", index=False)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Carga datos de cartera PF y genera gráficos de market share."
    )
    parser.add_argument(
        "--cutoff",
        default=CUTOFF_DEFAULT,
        help="Fecha de corte en formato YYYY-MM-DD (default: %(default)s)",
    )
    parser.add_argument(
        "--skip-load",
        action="store_true",
        help="No volver a cargar los Excel a DuckDB (usa la tabla existente)",
    )
    parser.add_argument(
        "--db-path",
        default=str(DB_PATH),
        help="Ruta al archivo DuckDB (default: %(default)s)",
    )
    return parser.parse_args()


def adjust_cutoff(conn: duckdb.DuckDBPyConnection, cutoff_str: str) -> str:
    """Ajusta el cutoff usando el rango disponible en la tabla cartera_pf."""
    res = conn.execute(
        """
        SELECT
            MIN(date_trunc('month', fecha)) AS min_month,
            MAX(date_trunc('month', fecha)) AS max_month
        FROM cartera_pf
        """
    ).fetchone()
    min_month, max_month = res
    if min_month is None or max_month is None:
        raise ValueError("cartera_pf está vacío; ejecuta el downloader primero.")

    cutoff_dt = pd.to_datetime(cutoff_str).to_period("M").to_timestamp()
    min_month = pd.to_datetime(min_month).to_period("M").to_timestamp()
    max_month = pd.to_datetime(max_month).to_period("M").to_timestamp()

    if cutoff_dt > max_month:
        print(
            f"Aviso: cutoff {cutoff_dt.date()} está fuera de rango. "
            f"Se ajusta a {max_month.date()}."
        )
        cutoff_dt = max_month
    if cutoff_dt < min_month:
        print(
            f"Aviso: cutoff {cutoff_dt.date()} anterior a los datos disponibles. "
            f"Se ajusta a {min_month.date()}."
        )
        cutoff_dt = min_month
    return cutoff_dt.strftime("%Y-%m-%d")


def main() -> None:
    args = parse_args()
    ensure_out_dir()
    db_path = Path(args.db_path)

    if not args.skip_load:
        df_cartera = build_cartera_pf()
        load_into_duckdb(df_cartera, db_path=db_path)

    conn = duckdb.connect(str(db_path))
    cutoff = adjust_cutoff(conn, args.cutoff)
    df_share = get_share_df(conn, cutoff)
    if df_share.empty:
        conn.close()
        raise ValueError(
            "No se encontraron registros en cartera_pf para la ventana solicitada. "
            "Verifica que existan datos suficientes."
        )

    df_compo = prepare_composition(df_share, cutoff)
    save_csvs(df_share, df_compo)
    share_path = plot_market_share(df_share, cutoff)
    compo_path = plot_composition_cutoff(df_share, cutoff)
    conn.close()

    print(f"Gráfico de market share guardado en: {share_path}")
    print(f"Gráfico de composición guardado en: {compo_path}")


if __name__ == "__main__":
    main()
